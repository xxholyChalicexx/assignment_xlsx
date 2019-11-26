from django.shortcuts import render
from django.http import HttpResponse
from .forms import locationForm
from django.conf import settings
from decouple import config
import random
import os
import requests
import xlrd
from openpyxl import load_workbook

CHAR = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789'
API_KEY = config('API_KEY')

# Create your views here.

def random_name():
    name = ""
    for i in range(10):
        name += random.choice(CHAR)
    return name

def name_filter(file_name):
    [name, extension] = file_name.split('.')
    random_string = random_name()
    name += "_" + random_string
    full_file_name = name + "." + extension
    return full_file_name

def get_long_lat(address):
    headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:12.0) Gecko/20100101 Firefox/12.0'}
    url = f'https://maps.googleapis.com/maps/api/geocode/json?address={address}&key={API_KEY}'
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        json_data = response.json()
        lat = json_data['results'][0]['geometry']['location']['lat']
        lng = json_data['results'][0]['geometry']['location']['lng']
        return (lat, lng)
    else:
        return "Uh-Oh Something wrong in the API"

def fetch_address(file):
    wb = xlrd.open_workbook(file)
    sheet = wb.sheet_by_index(0)
    row_num = sheet.nrows
    address_list = []
    for i in range(row_num):
        address_list.append(sheet.cell_value(i,0))
    return address_list

def write_address(file,long_lat_list):
    wb = load_workbook(file)
    sheet = wb.active
    row_len = sheet.max_row
    print(row_len)
    print(long_lat_list)
    for i in range(1, len(long_lat_list) + 1):
        sheet.cell(row = i, column=2).value = f'Latitude:{long_lat_list[i-1][0]}'
        sheet.cell(row = i, column=3).value = f'Longitude:{long_lat_list[i-1][1]}'
        i += 1
    wb.save(file)


def home(request):
    if request.method == 'POST':
        form = locationForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_name = name_filter(request.FILES['file'].name)
            if (file == '' or '.xlsx' not in file_name): # incase the file is not uploaded or if it is not an execl sheet
                return HttpResponse("No file uploaded Or wrong file format")
            file_path = os.path.join(settings.FILES_DIR, file_name)
            print(file_path)
            with open(file_path, 'wb+') as fp:
                for chunk in request.FILES['file']:
                    fp.write(chunk)
            address_list = fetch_address(file_path)#fetching all the address from a excel sheet to python list
            print(address_list)
            long = 0
            lat = 0
            long_lat_list = []
            for i in address_list:
                long,lat = get_long_lat(i)# using api for getting the latitude and longitude
                long_lat_list.append((long,lat))
            write_address(file_path, long_lat_list)#writing the latitude and longitude of the respetive address
            return HttpResponse(f'<a href="media\Files\{file_name}" download>File Download</a>')

    else:
        form = locationForm
    return render(request, 'app/file.html',{'form':form})